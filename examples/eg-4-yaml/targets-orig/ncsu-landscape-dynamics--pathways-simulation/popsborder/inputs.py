# Simulation of contaminated consignments and their inspections
# Copyright (C) 2018-2022 Vaclav Petras and others (see below)

# This program is free software; you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation; either version 2 of the License, or (at your option) any later
# version.

# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or
# FITNESS FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.

# You should have received a copy of the GNU General Public License along with
# this program; if not, see https://www.gnu.org/licenses/gpl-2.0.html


"""Inputs, especially loading of configuration

.. codeauthor:: Vaclav Petras <wenzeslaus gmail com>
"""

import copy
import json
import math
import sys
import types
from collections.abc import Iterable, Mapping
from pathlib import Path


def text_to_value(arg):
    """Convert text to int, float, or interpret it as JSON if possible

    If the argument is not string, it is returned as is.
    If conversion is not possible, the original text is returned.
    None is returned for both None and an empty string.
    """
    # This is a short function with return statements only.
    # pylint: disable=too-many-return-statements
    if not isinstance(arg, str):
        return arg
    if not arg:
        return None
    try:
        return int(arg)
    except ValueError:
        try:
            return float(arg)
        except ValueError:
            # Supports JSON booleans, Python-like booleans, and all YAML booleans.
            # We do not support old YAML booleans.
            if arg in ["true", "True", "TRUE"]:
                return True
            if arg in ["false", "False", "FALSE"]:
                return False
            try:
                return json.loads(arg)
            except json.JSONDecodeError:
                # add yaml.safe_load here?
                # Return the original value.
                return arg


def update_nested_dict_by_dict(dictionary, update):
    """Recursively update nested dictionary by anther nested dictionary"""
    for key, value in update.items():
        if isinstance(value, Mapping):
            dictionary[key] = update_nested_dict_by_dict(dictionary.get(key, {}), value)
        else:
            dictionary[key] = value
    return dictionary


def update_nested_dict_by_item(dictionary, keys, value):
    """Update nested dictionary by a nested keys-value pair

    An item is a list of keys to navigate the nested dictionary and a value to place
    in the given position.

    When a key can be represented as an int, it is used as a list index.
    List must already exists in the given size or the only index used must be 0.

    Floating point keys are not supported.
    """
    if len(keys) == 1:
        key = keys[0]
        try:
            key = int(key)
        except ValueError:
            pass
        dictionary[key] = value
    else:
        if keys[0] not in dictionary:
            try:
                # Test if the next key is an integer and thus index in a list.
                int(keys[1])
                # In case it is a list, we require keys are items are filled in order.
                dictionary[keys[0]] = [None]
            except ValueError:
                dictionary[keys[0]] = {}
        update_nested_dict_by_item(dictionary[keys[0]], keys[1:], value)


def record_to_nested_dictionary(record):
    """Convert dictionary with key/subkey/subsubkey keys into a nested dictionary"""
    out = {}
    for path, value in record.items():
        try:
            keys = path.split("/")
        except AttributeError as error:
            raise ValueError(
                f"Record items need to be strings, not {path} ({type(path).__name__})"
            ) from error
        update_nested_dict_by_item(out, keys, value)
    return out


def update_config(config, record):
    """Update config dictionary by a dictionary with key/subkey/subsubkey keys"""
    config = copy.deepcopy(config)
    update = record_to_nested_dictionary(record)
    update_nested_dict_by_dict(config, update)
    return config


def load_configuration_yaml_from_text(text):
    """Return configuration dictionary from YAML in a string"""
    import yaml  # pylint: disable=import-outside-toplevel

    if hasattr(yaml, "full_load"):
        return yaml.full_load(text)
    return yaml.load(text)  # pylint: disable=no-value-for-parameter


def load_one_configuration(filename, sheet=None, key_column=None, value_column=None):
    """Get the configuration from a JSON or YAML file

    The format is decided based on the file extension.
    It uses full_load() (FullLoader) to read YAML.

    The parameter can be a string or a path object (path-like object).

    If the *filename* contains `::`, anything after the last `::` is considered
    parameters determining where in the spreadsheet or table are the relevant
    columns. The format is multiple key-value pairs with key and value separated by
    `=`, `:`, or `: ` and individual pairs separated by `,`.
    The same information can be passed directly as function parameters.
    If both are provided, function parameters take precedence.
    """
    if isinstance(filename, Iterable) and not isinstance(filename, str):
        return record_to_nested_dictionary(filename)

    filename_str = str(filename)
    if "::" in filename_str:
        filename, info = filename_str.rsplit("::", maxsplit=1)
        info = table_info_from_text(info)
    else:
        info = table_info_from_text("")
    filename = Path(filename)
    if sheet:
        info.sheet = sheet
    if key_column:
        info.key_column = key_column
    if value_column:
        info.value_column = value_column

    if str(filename).endswith(".json"):
        return json.load(open(filename))
    elif str(filename).endswith(".yaml") or str(filename).endswith(".yml"):
        import yaml  # pylint: disable=import-outside-toplevel

        if hasattr(yaml, "full_load"):
            return yaml.full_load(open(filename))
        return yaml.load(open(filename))  # pylint: disable=no-value-for-parameter
    elif filename.suffix.lower() in [".csv", ".xlsx", ".ods"]:
        return load_config_table(
            filename,
            sheet=info.sheet,
            key_column=info.key_column,
            value_column=info.value_column,
        )
    else:
        sys.exit("Unknown file extension (file: {filename})")


def resolve_included_files(dictionary: dict, base_file_name=None):
    """Replace links to files by the file content (materialize included files)"""
    for key, value in dictionary.items():
        if isinstance(value, dict):
            if len(value) == 1 and "include_file" in value:
                value = value["include_file"]
                nested_file = value["file_name"]
                if base_file_name:
                    parent = Path(base_file_name).parent
                    nested_file = parent / nested_file
                nested_sheet = value.get("sheet")
                nested_key_column = value.get("key_column")
                nested_value_column = value.get("value_column")
                file_format = value.get("file_format")
                if file_format == "list":
                    values = load_scenario_table(nested_file)
                    new_value = []
                    for item in values:
                        new_value.append(record_to_nested_dictionary(item))
                else:
                    new_value = load_one_configuration(
                        nested_file,
                        sheet=nested_sheet,
                        key_column=nested_key_column,
                        value_column=nested_value_column,
                    )
                dictionary[key] = new_value
            else:
                resolve_included_files(value, base_file_name)


def load_configuration(filename, sheet=None, key_column=None, value_column=None):
    """Get the configuration from a JSON or YAML file

    The format is decided based on the file extension.
    It uses full_load() (FullLoader) to read YAML.

    The parameter can be a string or a path object (path-like object).

    If the *filename* contains `::`, anything after the last `::` is considered
    parameters determining where in the spreadsheet or table are the relevant
    columns. The format is multiple key-value pairs with key and value separated by
    `=`, `:`, or `: ` and individual pairs separated by `,`.
    The same information can be passed directly as function parameters.
    If both are provided, function parameters take precedence.

    Any file specified under the `include_file` key is included.
    """
    config = load_one_configuration(
        filename, sheet=sheet, key_column=key_column, value_column=value_column
    )
    resolve_included_files(config, base_file_name=filename)
    return config


def table_info_from_text(text, sheet=None, key_column=None, value_column=None):
    """Convert comma-separated list of key-value pairs to info about table

    Items are separated by comma. Key and value can be separated by
    `=` or `:` and also by `: ` and ` = ` because spaces surrounding
    key and value are removed.

    If there is no key-value pair, the whole value is used as a
    value for key_column.

    Function parameters are used as default values.

    Accepted keys are 'sheet', 'key_column', and 'value_column'.
    """
    info = types.SimpleNamespace(
        sheet=sheet, key_column=key_column, value_column=value_column
    )
    if not text:
        return info
    items = text.split(",")
    for item in items:
        key = None
        for separator in ["=", ":"]:
            if separator in item:
                key, value = item.split(separator, maxsplit=1)
                key = key.strip()
                value = value.strip()
                break
        if not key:
            info.key_column = item
        elif key in ["sheet", "key_column", "value_column"]:
            setattr(info, key, value)
        else:
            raise ValueError(f"Unknown key '{key}' in table info specification")
    return info


def load_config_table(filename, sheet=None, key_column=None, value_column=None):
    """Load a CSV file into a list of dictionaries

    Values which can be converted into int or float are converted. Cells which can be
    parsed as JSON, will be loaded into Python data structures (dicts, lists, etc.).

    A whole file is read and loaded into memory unlike with the ``csv.reader()``
    function.
    """
    if Path(filename).suffix.lower() not in [".csv", ".ods"]:
        return load_config_xlsx(
            filename, sheet=sheet, key_column=key_column, value_column=value_column
        )

    if Path(filename).suffix.lower() == ".ods":
        return load_config_ods(
            filename, sheet=sheet, key_column=key_column, value_column=value_column
        )

    return load_config_csv(filename, key_column=key_column, value_column=value_column)


def column_from_string(arg, default, fallback):
    """Return zero-based column index created from a column specification.

    Values convertable to integer are converted and considered to be one-based
    column index and thus further converted to zero-based index.

    For values not convertable to integer, fallback function is called and its
    result is returned. This allows for a backend-specific column conversion function
    to be used to resolve the complex cases.

    For None and empty string, the default value is returned.
    """
    if not arg:
        return default
    else:
        try:
            return int(arg) - 1
        except ValueError:
            return fallback(arg)


def validate_key(arg):
    """Return a validated key or None if it cannot be validated.

    Accepts any value which is supposed to be a configuration key
    and returns that value ready to be used as a key, i.e., a string
    without any extra whitespace around it (in case of user-input errors
    in a spreadsheet).

    For an empty string, NaN (not a number), and None returns None,
    i.e., the key is invalid and the associated value in a key-value pair (if any)
    should be ignored. Additionally, it returns None also for keys containing
    a space (after being stripped from leading and trainling spaces). These are
    assumed to be headings inside a spreadsheet or other values to be ignored
    because valid keys don't have spaces in them.
    """
    if isinstance(arg, str):
        # Allow for whitespace around the key.
        arg = arg.strip()
        if not arg or " " in arg:
            return None
        return arg
    if arg is None:
        return None
    if isinstance(arg, float) and math.isnan(arg):
        # NaN is None in our context since we don't use NaN in configuration
        # and Pandas reader defaults to NaN for no-data values.
        return None
    # All keys should be strings, so convert numbers and anything else to strings.
    return str(arg)


def load_config_csv(filename, key_column=None, value_column=None):
    """Read configuration from a CSV table.

    The key_column and value_column parameters can be one-based column indices or
    spreadsheet-like column letters in range A-Z.

    Rows without keys or with invalid keys are ignored. See :func:`validate_key`
    for details.

    Values which can be converted from string to other types are converted
    automatically. See :func:`text_to_value` for details.
    """
    table = {}
    with open(filename) as file:
        # pylint: disable=import-outside-toplevel
        import csv

        key_column = column_from_string(
            key_column, default=0, fallback=lambda x: ord(x) - ord("A")
        )
        value_column = column_from_string(
            value_column, default=key_column + 1, fallback=lambda x: ord(x) - ord("A")
        )

        for row in csv.reader(file):
            key = validate_key(row[key_column])
            if key:
                value = text_to_value(row[value_column])
                table[key] = value

    return record_to_nested_dictionary(table)


def load_config_xlsx(filename, sheet=None, key_column=None, value_column=None):
    """Read configuration from a XLSX table.

    The sheet parameter is name of the sheet within the given spreadsheet file.

    The key_column and value_column parameters can be one-based column indices or
    spreadsheet-like column letters.

    See :func:`load_config_csv` for handling of keys and values.
    """
    table = {}

    # pylint: disable=import-outside-toplevel
    import openpyxl
    from openpyxl.utils import column_index_from_string

    key_column = column_from_string(
        key_column, default=0, fallback=lambda x: column_index_from_string(x) - 1
    )
    value_column = column_from_string(
        value_column,
        default=key_column + 1,
        fallback=lambda x: column_index_from_string(x) - 1,
    )

    import warnings

    with warnings.catch_warnings():
        # We want to use validation functions in the spreadsheet,
        # but it does not matter whether they are supported by the reader here.
        warnings.filterwarnings(
            "ignore", message="Data Validation extension is not supported"
        )
        workbook = None
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)
            if sheet:
                sheet = workbook[sheet]
            else:
                sheet = workbook.active
            # Read rows excluding the header.
            for row in sheet.iter_rows(values_only=True):
                key = validate_key(row[key_column])
                if key:
                    # Consider only rows with filled key to allow
                    # for empty rows for formatting purposes.
                    # Additionally, ignore rows where key cell contains
                    # spaces (this allows for a header without threating
                    # first row differently).
                    value = text_to_value(row[value_column])
                    table[key] = value
        finally:
            # Read-only mode requires an explicit close and
            # the workbook object is not a context manager.
            if workbook:
                workbook.close()
        return record_to_nested_dictionary(table)


def load_config_ods(filename, sheet=None, key_column=None, value_column=None):
    """Read configuration from a ODS table.

    The sheet parameter is name of the sheet within the given spreadsheet file.

    The key_column and value_column parameters can be one-based column indices or
    spreadsheet-like column letters. If the value column comes before the key
    (parameter name) column in the file, only columns in the range A-Z are supported
    when specified using letters (specifying using one-based indices works
    in any case).

    See :func:`load_config_csv` for handling of keys and values.
    """
    table = {}

    # pylint: disable=import-outside-toplevel
    import pandas

    sheet = sheet if sheet else 0

    key_column = column_from_string(key_column, default=0, fallback=lambda x: x)
    value_column = column_from_string(value_column, default=1, fallback=lambda x: x)

    # The columns are always returned in the order as in the table, not in
    # the order specified in the parameter, so we need to know the relative order
    # of the two columns.
    flip_order = False
    # Assuming both are of the same type.
    if isinstance(key_column, int) and isinstance(value_column, int):
        cols = [key_column, value_column]
        if key_column > value_column:
            flip_order = True
    else:
        cols = f"{key_column},{value_column}"
        # Multi-letter columns which are in opposite order are not handled.
        if (
            len(key_column) == 1
            and len(value_column) == 1
            and ord(key_column) > ord(value_column)
        ):
            flip_order = True

    if flip_order:
        key_column_index = 2
        value_column_index = 1
    else:
        key_column_index = 1
        value_column_index = 2

    data = pandas.read_excel(
        Path(filename), sheet_name=sheet, header=None, usecols=cols
    )
    table = {}
    for row in data.itertuples():
        key = validate_key(row[key_column_index])
        if key:
            value = text_to_value(row[value_column_index])
            if isinstance(value, float) and math.isnan(value):
                value = None
            table[key] = value
    return record_to_nested_dictionary(table)


def add_dict_config_to_table(table, value, keys=None):
    """Add a nested dictionary to a table represented by a mapping

    This is meant for internal use, if possible, use :func:`dict_config_to_table`.
    """
    if keys is None:
        keys = []
    if isinstance(value, Mapping):
        for nested_key, nested_value in value.items():
            new_keys = keys.copy()
            new_keys.append(nested_key)
            add_dict_config_to_table(table, value=nested_value, keys=new_keys)
    else:
        table["/".join(keys)] = value


def dict_config_to_table(value):
    """Convert a nested dictionary to a table represented by a mapping"""
    table = {}
    add_dict_config_to_table(table, value)
    return table


def print_table_config(config, file=None):
    """Print a table represented by a mapping

    Generates key-value pairs separated by a pipe.
    It does not quote or espace the separator in values.
    """
    for key, value in config.items():
        if isinstance(value, Iterable) and not isinstance(value, str):
            value = json.dumps(value)
        print(f"{key}|{value}", file=file)


def load_scenario_table(filename):
    """Load a CSV file into a list of dictionaries

    Values which can be converted into int or float are converted. Cells which can be
    parsed as JSON, will be loaded into Python data structures (dicts, lists, etc.).

    A whole file is read and loaded into memory unlike with the ``csv.reader()``
    function.
    """
    # pylint: disable=import-outside-toplevel
    table = []

    # Read spreadsheet formats
    if Path(filename).suffix.lower() != ".csv":
        import openpyxl

        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)
            sheet = workbook.active
            # Get header.
            header = [cell.value for cell in sheet[1]]
            # Read rows excluding the header.
            for old_row in sheet.iter_rows(min_row=2):
                new_row = {}
                for key, cell in zip(header, old_row):
                    new_row[key] = text_to_value(cell.value)
                table.append(new_row)
        finally:
            # Read-only mode requires an explicit close and
            # the workbook object is not a context manager.
            workbook.close()
        return table

    # Read as CSV
    with open(filename) as file:
        import csv

        for row in csv.DictReader(file):
            for key, value in row.items():
                row[key] = text_to_value(value)
            table.append(row)
    return table


def load_cfrp_schedule(filename, date_format=None):
    """Load cut flower release program (CFRP) schedule

    Supports one CSV file with all combinations of date, origin, and commodity (flower)
    row by row with columns date, origin_nm, and commodity (caseinsensitive).
    Duplicate records don't appear in the resulting data structure.

    The values in the date column are parsed according to *date_format*.

    Returns a dictionary with keys being tuples of commodity and origin and values
    being a set of dates.
    """
    if not date_format:
        date_format = "%Y-%m-%d"
    schedule = {}
    # Read as CSV
    with open(filename) as file:
        # Import file- or format-specific items only when need.
        # pylint: disable=import-outside-toplevel
        import csv
        from datetime import datetime

        for row in csv.DictReader(file):
            for key, value in row.items():
                if key.lower() == "date":
                    date = datetime.strptime(value, date_format).date()
                elif key.lower() == "origin_nm":
                    origin = value
                elif key.lower() == "commodity":
                    commodity = value
            combo = (commodity, origin)
            if combo not in schedule:
                # Using set to ensure we have no duplicate dates.
                schedule[combo] = set()
            schedule[combo].add(date)
    return schedule


def load_skip_lot_consignment_records(filename, tracked_properties):
    """Load records associating consignment with skip lot compliance levels.

    Only the *tracked_properties* are considered and are assumed to uniquely
    distinguish consignment records with distinct compliance levels.
    If the level can be converted to number, it is converted.
    """
    records = {}
    # Read as CSV
    with open(filename) as file:
        # Import file- or format-specific items only when need.
        # pylint: disable=import-outside-toplevel
        import csv

        for row in csv.DictReader(file):
            combo = []
            for tracked_property in tracked_properties:
                combo.append(row[tracked_property])
            level = row["compliance_level"]
            records[tuple(combo)] = text_to_value(level)
    return records
